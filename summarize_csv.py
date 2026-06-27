import os
import glob
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ============================================================
# 設定
# ============================================================
ANSWER_DIR = "csv"                        # 回答CSVを入れるフォルダ
OUTPUT_DETAIL_CSV = "回答一覧_集計用.csv"
OUTPUT_EXCEL = "審査結果.xlsx"
LOG_FILE = "log_summarize.txt"

# 投票結果の意味
# N: 未審査、J: 審査済み・推薦なし、R: 審査済み・推薦あり
VALID_RESULTS = {"N", "J", "R", ""}

# ============================================================
# CSV読み込み
# ============================================================
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parent
ANSWER_DIR = BASE_DIR / "csv"


def read_csv_safely(path):
    """日本語CSVでよく使う文字コードを順に試す。"""
    for enc in ["utf-8-sig", "utf-8", "cp932"]:
        try:
            return pd.read_csv(path, header=None, encoding=enc)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, header=None)


def get_answer_files():
    files = sorted(ANSWER_DIR.glob("*.csv"))
    if not files:
        raise FileNotFoundError(f"{ANSWER_DIR} フォルダ内にCSVファイルがありません。")
    return files


def check_format(df, path):
    """想定形式か確認する。"""
    if df.shape[0] < 3 or df.shape[1] < 2:
        raise ValueError(f"CSVの行数または列数が足りません: {path}")

    if str(df.iloc[0, 0]).strip() != "氏名":
        raise ValueError(f"1行目1列目が『氏名』ではありません: {path}")
    if str(df.iloc[0, 2]).strip() != "博士取得年":
        raise ValueError(f"1行目3列目が『博士取得年』ではありません: {path}")
    if str(df.iloc[1, 0]).strip() != "発表者名":
        raise ValueError(f"2行目1列目が『発表者名』ではありません: {path}")
    if str(df.iloc[1, 1]).strip() != "投票結果":
        raise ValueError(f"2行目2列目が『投票結果』ではありません: {path}")


def get_voter_label(df, path):
    """氏名_博士取得年 の形で回答者ラベルを作る。"""
    name = str(df.iloc[0, 1]).strip()
    phd_year = str(df.iloc[0, 3]).strip().replace(".0", "") if df.shape[1] >= 4 else ""

    if not name or name.lower() == "nan":
        name = os.path.splitext(os.path.basename(path))[0]

    return f"{name}_{phd_year}" if phd_year and phd_year.lower() != "nan" else name


def read_one_answer(path):
    """1つの回答CSVを読み込む。"""
    df = read_csv_safely(path)
    check_format(df, path)

    voter_label = get_voter_label(df, path)

    answer_df = df.iloc[2:, :2].copy()
    answer_df.columns = ["発表番号_発表者名", "投票結果"]
    answer_df = answer_df.dropna(subset=["発表番号_発表者名"])

    answer_df["発表番号_発表者名"] = answer_df["発表番号_発表者名"].astype(str).str.strip()
    answer_df["投票結果"] = answer_df["投票結果"].fillna("").astype(str).str.strip().str.upper()

    return voter_label, answer_df


# ============================================================
# 集計用の変換
# ============================================================
def status_to_counts(status):
    """N/J/R/エラーを、審査数・投票数に変換する。"""
    if status == "R":
        return 1, 1
    if status == "J":
        return 1, 0
    return 0, 0


def main():
    answer_files = get_answer_files()
    log_lines = ["== Alerts =="]

    # --------------------------------------------------------
    # 1. すべてのCSVを読み込み、発表一覧を作る
    # --------------------------------------------------------
    all_answers = []
    presentation_labels = []

    for path in answer_files:
        voter_label, answer_df = read_one_answer(path)
        all_answers.append((path, voter_label, answer_df))

        for label in answer_df["発表番号_発表者名"]:
            if label not in presentation_labels:
                presentation_labels.append(label)

    if not presentation_labels:
        raise ValueError("発表一覧を作成できませんでした。")

    # --------------------------------------------------------
    # 2. 回答一覧を作る
    # --------------------------------------------------------
    detail_df = pd.DataFrame({"発表番号_発表者名": presentation_labels})

    sum_inspects = [0] * len(presentation_labels)
    sum_votes = [0] * len(presentation_labels)
    has_error = [False] * len(presentation_labels)

    for path, voter_label, answer_df in all_answers:
        result_map = dict(zip(answer_df["発表番号_発表者名"], answer_df["投票結果"]))
        statuses = []

        for i, label in enumerate(presentation_labels):
            status = result_map.get(label, "N")

            if status not in VALID_RESULTS:
                log_lines.append(f"{path}: {label}: 投票結果が N/J/R ではありません: {status}")
                status = "エラー"
                has_error[i] = True

            inspect_count, vote_count = status_to_counts(status)
            sum_inspects[i] += inspect_count
            sum_votes[i] += vote_count
            statuses.append(status if status else "N")

        detail_df[voter_label] = statuses

    detail_df.to_csv(BASE_DIR / OUTPUT_DETAIL_CSV, index=False, encoding="utf-8-sig")

    # --------------------------------------------------------
    # 3. 集計表を作る
    # --------------------------------------------------------
    vote_ratios = []
    for label, inspect_count, vote_count in zip(presentation_labels, sum_inspects, sum_votes):
        if inspect_count < 3:
            vote_ratios.append(0)
            log_lines.append(f"{label}: 審査数が3未満です。")
        else:
            vote_ratios.append(vote_count / inspect_count)

    summary_df = pd.DataFrame({
        "発表番号_発表者名": presentation_labels,
        "審査数合計": sum_inspects,
        "投票数合計": sum_votes,
        "投票数/審査数": vote_ratios,
        "入力エラー": ["あり" if v else "" for v in has_error],
    })

    summary_sorted_df = summary_df.sort_values(
        ["投票数/審査数", "投票数合計", "審査数合計"],
        ascending=[False, False, False]
    )

    with pd.ExcelWriter(BASE_DIR / OUTPUT_EXCEL, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="集計", index=False)
        summary_sorted_df.to_excel(writer, sheet_name="投票率順", index=False)
        detail_df.to_excel(writer, sheet_name="回答一覧", index=False)

    # --------------------------------------------------------
    # 4. Excelの見た目を整える
    # --------------------------------------------------------
    wb = load_workbook(BASE_DIR / OUTPUT_EXCEL)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    warning_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    error_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)

    for sheet_name in ["集計", "投票率順"]:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            inspect_count = ws.cell(row=row, column=2).value
            error_value = str(ws.cell(row=row, column=5).value or "").strip()

            if inspect_count < 3:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = warning_fill

            if error_value == "あり":
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = error_fill

            ws.cell(row=row, column=4).number_format = "0.0%"

    ws = wb["回答一覧"]
    for row in ws.iter_rows(min_row=2):
        for cell in row[1:]:
            if cell.value == "エラー":
                cell.fill = error_fill

    wb.save(BASE_DIR / OUTPUT_EXCEL)

    # --------------------------------------------------------
    # 5. ログ出力
    # --------------------------------------------------------
    with open(BASE_DIR / LOG_FILE, "w", encoding="utf-8") as fp:
        fp.write("\n".join(log_lines) + "\n")

    print(f"Done: {OUTPUT_DETAIL_CSV}, {OUTPUT_EXCEL}, {LOG_FILE}")


if __name__ == "__main__":
    main()
